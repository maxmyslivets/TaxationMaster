import os
import pickle
import shutil
import time
import traceback
from pathlib import Path

from PySide6.QtCore import Qt
from docx import Document as DocxDocument
from openpyxl import load_workbook

from PySide6 import QtCore
from PySide6.QtGui import QAction, QColor, QBrush
from PySide6.QtWidgets import QFileDialog, QTreeWidgetItem, QTreeWidget, QMenu, QTableView, QTableWidget, \
    QTableWidgetItem, QComboBox, QTabWidget

from source.model.project import ProjectData, Project
from source.view.configuration_import import ConfigurationImport
from source.view.custom_widgets import EditableComboBox, CustomTabWidget
from utils.convert import oda_converter


class Interface:
    project_data: ProjectData

    def __init__(self, model, view):
        self.model = model
        self.view = view

    ###################################################################################################################
    # Меню Проект
    ###################################################################################################################

    def save_as_project(self) -> None:
        save_as = QFileDialog()
        save_as.setDefaultSuffix(self.model.config.extension)
        _project_path, _ = save_as.getSaveFileName(parent=self.view.main_window, caption="Сохранить как...", dir='/',
                                                   filter=f"Taxation tool project (*{self.model.config.extension})")
        if _project_path:
            try:
                project_path = Path(_project_path)
                if project_path.exists():
                    os.remove(project_path)
                    shutil.rmtree(project_path.parent / project_path.stem)
                elif project_path == self.project_data.path:
                    self.save_project()
                self.project_data.path = project_path
                self.model.project.is_saved = True
                with open(self.project_data.path, 'wb') as file:
                    pickle.dump(self.model.project, file)

                self.clear_temp_project()
                self.update_interface()

                self.view.main_window.log(f"[DEBUG]\tПроект `{self.project_data.path}` успешно сохранен.")
            except Exception:
                self.view.main_window.log(f"[ERROR]\tНе удалось сохранить проект в `{_project_path}`."
                                          f"\n{traceback.format_exc()}")

    def save_project(self) -> None:
        try:
            with open(self.project_data.path, 'wb') as file:
                pickle.dump(self.model.project, file)
            self.model.project.is_saved = True
            self.view.main_window.log(f"[DEBUG]\tПроект `{self.project_data.path}` успешно сохранен.")
        except Exception:
            self.view.main_window.log(f"[ERROR]\tНе удалось сохранить проект `{self.project_data.path}`."
                                      f"\n{traceback.format_exc()}")

    def create_new_project(self) -> None:

        # проверка на сохранение проекта перед созданием/открытием нового
        if not self.model.project.is_saved and self.view.ConfirmCloseUnsavedProject().result_no:
            return

        try:
            self.clear_temp_project()
            self.model.clear_project()

            # создание необходимых для работы директорий во временном каталоге
            try:
                os.makedirs(self.model.config.temp_path)
                os.makedirs(self.model.config.temp_path_convert_input)
                os.makedirs(self.model.config.temp_path_convert_output)
            except Exception as e:
                self.view.main_window.log(f"[ERROR]\tОшибка создания временной директории "
                                          f"`{self.model.config.temp_path}`.\n{str(e)}")
                shutil.rmtree(self.model.config.temp_path, ignore_errors=True)
                return

            # создание проекта во временном каталоге
            self.project_data = ProjectData(self.model.config.temp_path / ("New project" + self.model.config.extension))
            self.model.project.is_saved = True

            self.update_interface()
            self.clear_project_manager()
            # self.view.main_window.table.setRowCount(0)
            # self.view.main_window.table.setColumnCount(0)

            self.view.main_window.log(f"[DEBUG]\tПроект `{self.project_data.path}` успешно создан.")

        except Exception:
            self.view.main_window.log(f"[ERROR]\tОшибка создания проекта во временной директории."
                                      f"\n{traceback.format_exc()}")

    def open_project(self) -> None:

        # проверка на сохранение проекта перед созданием/открытием нового
        if not self.model.project.is_saved and self.view.ConfirmCloseUnsavedProject().result_no:
            return

        open_dialog = QFileDialog()
        open_dialog.setDefaultSuffix(self.model.config.extension)
        _project_path, _ = open_dialog.getOpenFileName(parent=self.view.main_window, caption="Открыть проект...", dir='/',
                                                       filter=f"Taxation tool project (*{self.model.config.extension})")
        if _project_path:
            try:
                project_path = Path(_project_path)
                if project_path.suffix != self.model.config.extension:
                    raise ValueError(f"Неверное расширение файла: `{self.model.config.extension}`. "
                                     f"Требуется `{self.model.project.suffix}`")
                with open(project_path, 'rb') as file:
                    self.model.project = pickle.load(file)

                self.model.project.is_saved = True

                self.clear_temp_project()

                self.view.main_window.log(f"[DEBUG]\tПроект `{self.project_data.path}` успешно открыт.")

                self.project_data = ProjectData(project_path)
                self.update_interface()
                self.view.main_window.table.setRowCount(0)
                self.view.main_window.table.setColumnCount(0)

                self.clear_project_manager()
                if "taxation_plan" in self.model.project.__dict__.keys():
                    self.set_taxation_plan_to_project_manager()
                if "taxation_list" in self.model.project.__dict__.keys():
                    self.set_taxation_list_to_project_manager()

            except Exception:
                self.view.main_window.log(f"[ERROR]\tНе удалось открыть проект `{_project_path}`."
                                          f"\n{traceback.format_exc()}")

    def clear_temp_project(self) -> None:
        self.view.main_window.log("[DEBUG]\tУдаление временных файлов.")
        if self.model.config.temp_path.exists():
            shutil.rmtree(self.model.config.temp_path)
        if self.model.config.temp_path_convert_input.exists():
            shutil.rmtree(self.model.config.temp_path_convert_input)
        if self.model.config.temp_path_convert_output.exists():
            shutil.rmtree(self.model.config.temp_path_convert_output)

    def import_taxation_plan(self) -> None:
        import_dialog = QFileDialog()
        dwg_path, _ = import_dialog.getOpenFileName(parent=self.view.main_window, caption="Импорт чертежа...", dir='/',
                                                    filter="Чертежи (*.dwg)")
        if dwg_path == "":
            return

        # конвертация dwg в dxf
        dwg_path_without_space = Path(dwg_path).name.replace(" ", "_")

        if self.model.config.temp_path_convert_input.exists():
            shutil.rmtree(self.model.config.temp_path_convert_input)
        os.makedirs(self.model.config.temp_path_convert_input, exist_ok=True)

        shutil.copyfile(Path(dwg_path), self.model.config.temp_path_convert_input / dwg_path_without_space)

        if self.model.config.temp_path_convert_output.exists():
            shutil.rmtree(self.model.config.temp_path_convert_output)
        os.makedirs(self.model.config.temp_path_convert_output, exist_ok=True)

        try:
            oda_converter(self.model.config.oda_converter_path, self.model.config.temp_path_convert_input,
                          self.model.config.temp_path_convert_output)
        except Exception:
            self.view.main_window.log(f"[ERROR]\tОшибка конвертации файлов."
                                      f"\n{traceback.format_exc()}")
            return

        timeout = self.model.config.timeout
        dxf_file = None
        while not dxf_file:
            try:
                dxf_file = self.model.config.temp_path_convert_output / next(
                    file for file in os.listdir(self.model.config.temp_path_convert_output))
            except StopIteration:
                time.sleep(1)
                timeout -= 1
            if timeout == 0:
                self.view.main_window.log(f"[ERROR]\tНе удалось конвертировать файл dxf. Превышено время ожидания.")
                break
        if timeout == 0:
            return

        timeout = self.model.config.timeout
        permission = False
        while not permission:
            try:
                permission = True
            except PermissionError:
                time.sleep(1)
                timeout -= 1
                if timeout == 0:
                    self.view.main_window.log(f"[ERROR]\tНе удалось импортировать файл dxf. Превышено время ожидания.")
                    break
        if timeout == 0:
            return

        taxation_plan = self.model.processing.create_taxation_plan(dxf_file,
                                                                   self.model.config.numbers_layers,
                                                                   self.model.config.lines_layers,
                                                                   self.model.config.contours_layers,
                                                                   self.model.config.min_distance)
        if taxation_plan is not None:
            self.model.project.taxation_plan = taxation_plan
            self.model.project.is_saved = False

        self.set_taxation_plan_to_project_manager()

        tab_widget: CustomTabWidget = self.view.main_window.tab_widget
        tab_widget.open_tab("Чертеж таксации")
        tab_widget.update_data_to_table(self.model.project.taxation_plan.table_data, "Чертеж таксации")

        self.show_in_table("Чертеж таксации")

        self.view.main_window.log(f"[DEBUG]\tЧертеж таксации успешно импортирован.")
        shutil.rmtree(self.model.config.temp_path_convert_output)

    def import_taxation_list(self) -> None:
        import_dialog = QFileDialog()
        doc_path, _ = import_dialog.getOpenFileName(parent=self.view.main_window,
                                                    caption="Импорт ведомости таксации...", dir='/',
                                                    filter="Документы Word/Excel (*.docx *.xlsx)")
        if doc_path == "":
            return

        list_of_tables = []

        if doc_path.endswith(".docx"):
            docx_document = DocxDocument(doc_path)
            for table in docx_document.tables:
                table_data = []
                for row in table.rows:
                    row_data = [cell.text.strip() for cell in row.cells]
                    table_data.append(row_data)
                list_of_tables.extend(table_data)

        elif doc_path.endswith(".xlsx") or doc_path.endswith(".xls"):
            xlsx_document = load_workbook(doc_path)
            sheet = xlsx_document[xlsx_document.sheetnames[0]]
            for row in sheet.iter_rows(values_only=True):
                row_data = [cell if cell is not None else "" for cell in row]
                list_of_tables.append(row_data)

        import_parameters = ConfigurationImport(list_of_tables).parameters

        if not import_parameters["is_import_first_row"]:
            list_of_tables.pop(0)

        edited_list_of_tables = []
        for row in list_of_tables:
            number, name, quantity, height, diameter, quality = row
            parts_idx_column = (
                (import_parameters["number"], number),
                (import_parameters["name"], name),
                (import_parameters["quantity"], quantity),
                (import_parameters["height"], height),
                (import_parameters["diameter"], diameter),
                (import_parameters["quality"], quality),
            )
            parts_idx_column = sorted(parts_idx_column, key=lambda x: x[0])
            edited_row = []
            for _, value in parts_idx_column:
                edited_row.append(value)
            edited_list_of_tables.append(edited_row)

        taxation_list = self.model.processing.create_taxation_list(edited_list_of_tables)
        if taxation_list is not None:
            self.model.project.taxation_list = taxation_list
            self.model.project.is_saved = False

        self.set_taxation_list_to_project_manager()

        tab_widget: CustomTabWidget = self.view.main_window.tab_widget
        tab_widget.open_tab("Ведомость таксации")
        tab_widget.update_data_to_table(self.model.project.taxation_list.table_data, "Ведомость таксации")

        self.show_in_table("Ведомость таксации")

        self.view.main_window.log(f"[DEBUG]\tВедомость таксации успешно импортирована.")

    def update_interface(self) -> None:
        self.view.main_window.setWindowTitle("Taxation Tool - " + self.project_data.name)
        self.view.main_window.log("[DEBUG]\tОбновление интерфейса.")

    ###################################################################################################################
    # Меню Обработка
    ###################################################################################################################

    # def preprocessing(self) -> None:
    #     self.model.processing.read_data_from_taxation_plan(self.model.config.numbers_layers,
    #                                                        self.model.config.lines_layers,
    #                                                        self.model.config.contours_layers,
    #                                                        self.model.config.zones_layers,
    #                                                        self.model.config.min_distance,
    #                                                        self.model.config.min_area)
    #     self.view.main_window.log("Файл чертежа таксации успешно обработан.")
    #     self.view.main_window.log(f"Количество точечных растений: {len(self.model.project.numbers)}")
    #     self.view.main_window.log(f"Количество полос и контуров растительности: {len(self.model.project.shapes)}")
    #     self.view.main_window.log(f"Зоны: {[name for _, name in self.model.project.zone_names.items()]}")
    #     self.model.processing.splitting_numbers()
    #     self.model.processing.calculate_intersects_shapes_in_zones()
    #
    #     split_numbers_in_zones_from_model = {zone_name: [] for _, zone_name in self.model.project.zone_names.items()}
    #     for zone_name in split_numbers_in_zones_from_model.keys():
    #         k_zone_name = next(k for k, v in self.model.project.zone_names.items() if v == zone_name)
    #         k_zone_list = self.model.project.zones_from_zone_names[k_zone_name]
    #         for k_zone in k_zone_list:
    #             for k_split_number, _k_zone_list in self.model.project.intersects_shapes_in_zones.items():
    #                 for _k_zone in _k_zone_list:
    #                     if _k_zone == k_zone:
    #                         split_numbers_in_zones_from_model[zone_name].append(self.model.project.split_numbers[k_split_number])
    #     for zone_name in split_numbers_in_zones_from_model.keys():
    #         self.view.main_window.log(f"Вхождения объектов в зону `{zone_name}`: "
    #                       f"{split_numbers_in_zones_from_model[zone_name]}")
    #     self.model.processing.splitting_shapes_in_zones()
    #     self.update_interface()
    #     self.show_table_from_taxation_plan()

    ###################################################################################################################
    # Управление виджетом менеджера проекта
    ###################################################################################################################

    def clear_project_manager(self) -> None:
        manager_project: QTreeWidget = self.view.main_window.tree_manager
        manager_project.clear()
        manager_project_taxation_plan = QTreeWidgetItem([f"Чертеж таксации"])
        manager_project.insertTopLevelItem(0, manager_project_taxation_plan)
        manager_project_taxation_list = QTreeWidgetItem([f"Ведомость таксации"])
        manager_project.insertTopLevelItem(1, manager_project_taxation_list)

    def set_taxation_plan_to_project_manager(self) -> None:
        manager_project: QTreeWidget = self.view.main_window.tree_manager
        try:
            manager_project_taxation_plan: QTreeWidgetItem = manager_project.findItems(
                "Чертеж таксации",
                QtCore.Qt.MatchStartsWith | QtCore.Qt.MatchRecursive)[0]
            index = manager_project.indexOfTopLevelItem(manager_project_taxation_plan)
            manager_project.takeTopLevelItem(index)
        except IndexError:
            pass
        manager_project_taxation_plan = QTreeWidgetItem([f"Чертеж таксации ({len(self.model.project.taxation_plan.numbers)})"])
        manager_project.insertTopLevelItem(0, manager_project_taxation_plan)

    def set_taxation_list_to_project_manager(self) -> None:
        manager_project: QTreeWidget = self.view.main_window.tree_manager
        try:
            manager_project_taxation_list: QTreeWidgetItem = manager_project.findItems(
                "Ведомость таксации",
                QtCore.Qt.MatchStartsWith | QtCore.Qt.MatchRecursive)[0]
            index = manager_project.indexOfTopLevelItem(manager_project_taxation_list)
            manager_project.takeTopLevelItem(index)
        except IndexError:
            pass
        manager_project_taxation_list = QTreeWidgetItem([f"Ведомость таксации "
                                                         f"({len(self.model.project.taxation_list.numbers)})"])
        manager_project.insertTopLevelItem(1, manager_project_taxation_list)

    def project_manager_double_clicked(self) -> None:
        manager_project: QTreeWidget = self.view.main_window.tree_manager
        item: QTreeWidgetItem = manager_project.selectedItems()[0]
        # Действия при нажатии на "Чертеж таксации"
        if item and item.parent() is None and item.text(0).startswith("Чертеж таксации")\
                and item.text(0) != "Чертеж таксации":
            self.show_in_table("Чертеж таксации")
        # Действия при нажатии на "Ведомость таксации"
        elif item and item.parent() is None and item.text(0).startswith("Ведомость таксации")\
                and item.text(0) != "Ведомость таксации":
            self.show_in_table("Ведомость таксации")

    def project_manager_context_menu(self, pos) -> None:
        manager_project: QTreeWidget = self.view.main_window.tree_manager
        item: QTreeWidgetItem = manager_project.itemAt(pos)
        menu = QMenu()
        # Действия при нажатии на "Чертеж таксации"
        if item and item.parent() is None and item.text(0).startswith("Чертеж таксации"):
            import_taxation_plan_action = QAction("Новый импорт")
            import_taxation_plan_action.triggered.connect(self.import_taxation_plan)
            menu.addAction(import_taxation_plan_action)
        elif item and item.parent() is None and item.text(0).startswith("Ведомость таксации"):
            import_taxation_list_action = QAction("Новый импорт")
            import_taxation_list_action.triggered.connect(self.import_taxation_list)
            menu.addAction(import_taxation_list_action)
        else:
            return
        menu.exec_(manager_project.viewport().mapToGlobal(pos))

    ###################################################################################################################
    # Управление виджетом таблицы
    ###################################################################################################################

    def show_in_table(self, tab_name: str) -> None:

        # TODO: Для реализации возможности внесения изменений в разделение номеров необходимо:
        #  1. при неудачном определении регулярного выражения подсвечивать строку красным
        #  2. добавить на строки таблицы контекстное меню "разделить строку & принудительная валидация".
        #  3. ячейки столбца "номер" представляют listBox с возможностью написания номера (проверка на написание только
        #     из имеющихся)
        #  4. при разделении строки вставлять строку ниже, подсветить желтым
        #  5. при нажатии "принудительная валидация" строку подсветить зеленым
        #     и добавить данные в split_numbers и number_from_split_number

        self.view.main_window.tab_widget.open_tab(tab_name)

    def table_context_menu(self, pos) -> None:
        table: QTableWidget = self.view.main_window.table
        # Определяем строку, где было вызвано меню
        index = table.indexAt(pos)
        if not index.isValid():
            return  # Игнорируем, если клик был вне таблицы

        row = index.row()

        # Создаем контекстное меню
        menu = QMenu(table)
        split_row_action = menu.addAction("Разделить строку")
        apply_action = menu.addAction("Подтвердить")

        # Обработка нажатия на пункт меню
        split_row_action.triggered.connect(lambda: self.split_table_row(row))
        apply_action.triggered.connect(lambda: self.apply_table_row(row))

        # Показываем меню
        menu.exec_(table.viewport().mapToGlobal(pos))

    def split_table_row(self, row):
        table: QTableWidget = self.view.main_window.table
        # Копируем данные из текущей строки
        copied_data = [table.item(row, col).text() if table.item(row, col) else "" for col in range(table.columnCount())]
        # Вставляем новую строку ниже текущей
        table.insertRow(row + 1)

        # Заполняем новую строку скопированными данными
        for col, data in enumerate(copied_data):
            new_item = QTableWidgetItem(data)
            new_item.setForeground(QBrush(QColor(255, 255, 0)))  # Устанавливаем желтый цвет текста
            table.setItem(row + 1, col, new_item)

        tab_widget = self.view.main_window.tabWidget
        if tab_widget.tabText(tab_widget.indexOf(tab_widget.currentWidget())) == "Чертеж таксации":
            combobox_number = EditableComboBox(
                self.model.project.taxation_plan.numbers.values(), table.cellWidget(row, 1).currentText())
            table.setCellWidget(row + 1, 1, combobox_number)


    def apply_table_row(self, row):
        pass




